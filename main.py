from PyQt5 import QtWidgets,QtSql
from Ui_mainwindow import  Ui_MainWindow
# from Ui_dialog import Ui_Dialog
# from PyQt5.QtCore import QTimer,Qt,QSize,QThread,pyqtSignal
# from PyQt5.QtGui import QColor
import sys
import time
from os import mkdir
import os
import logging
import pymysql
from openpyxl import Workbook
from  openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter

from dialog import MyDialog
import json

class MyWindowShow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MyWindowShow, self).__init__()
        self.setupUi(self)
        # self.layout=QtWidgets.QGridLayout()
        self.setWindowTitle("数据库查询 V1.0")
        self.log_init()
        self.Button_connect.clicked.connect(self.mysql_connect)
        self.Button_export.clicked.connect(self.export_excel)
        self.cBox_pc_filter.currentIndexChanged.connect(self.select_pc)
        if not(os.path.exists("config.json")):
            f = open("config.json",'w')
            f.write('''{
"host": "101.132.158.171",
"user": "test_read",
"port": 3306,
"passwd": "Xj2021.",
"db": "xjlcdbnew",
"default":"True"} ''')
            f.close()
        try:
            f = open("config.json")
            self.cfg = json.load(f)
            f.close()
        except OSError as reason:
            logging.info(str(reason))
            f.close()
        self.child_dialog = MyDialog(self.cfg)
        # 定义子窗口
        self.actionset.triggered.connect(self.child_dialog.show)
        #绑定QT信号槽
        self.child_dialog.signal_data.connect(self.get_database_config)
        self.Button_exportdata.clicked.connect(self.export_data)
        self.Button_file = QtWidgets.QPushButton(self.centralWidget)
        self.Button_file.setText("打开文件")
        self.statusBar.addPermanentWidget(self.Button_file)
        self.Button_file.clicked.connect(self.open_file)
        self.Button_file.setHidden(True)

    def mysql_connect(self,):
        logging.info("mysql_connect()...")
        if self.Button_connect.text()=="连接":
            try:
                self.conn = pymysql.connect(
                            host = self.cfg["host"],
                            user = self.cfg["user"],
                            password = self.cfg["passwd"],
                            port = self.cfg["port"],
                            database = self.cfg["db"],
                            charset = "utf8")
                self.cursor_Xj=self.conn.cursor()
                self.statusBar.showMessage("主机已连接%s" %self.cfg["host"])
                logging.info("主机已连接%s" %self.cfg["host"])
                self.setWindowTitle("数据库查询 V1.0"+'___'+self.cfg['db'])
                self.Button_connect.setText("断开连接")
                # 获取所有存在的批次，添加到批次下拉框中
                self.get_pc()
            except OSError as reason:
                self.statusBar.showMessage("服务器连接失败..."+str(reason))
                logging.info("服务器连接失败...%s" %str(reason))
        else:
            self.conn.close()
            logging.info("已经断开连接%s" %self.cfg["host"])
            self.statusBar.showMessage("已经断开连接%s" %self.cfg["host"])
            self.setWindowTitle("数据库查询 V1.0")
            self.Button_connect.setText("连接")
            self.Button_file.setHidden(True)

    def get_pc(self):
        """
        get_pc:获取所有生产批次，添加到下拉框界面中...
        """
        logging.info("正在执行'get_pc'...")
        sql="""select pc from t_pc ORDER BY intime desc"""
        self.cursor_Xj.execute(sql)
        self.pc_list=list(self.cursor_Xj.fetchall())
        logging.info(self.pc_list)
        self.cBox_pc_filter.clear()
        self.cBox_pc_filter.addItem("请选择批次")
        for each in self.pc_list:
            self.cBox_pc_filter.addItems(each)
        

    def export_excel(self,):
        logging.info("export_excel()...")
        mid_start = self.Edit_MidStart.text()
        mid_end = self.Edit_MidEnd.text()
        if(len(mid_start)!=22 or len(mid_end)!=22):
            self.statusBar.showMessage("模块ID输入有误...")
            logging.info("模块ID输入有误...")
        else:
            sql = '''SELECT modulid as'模块ID',LEFT(icid,48) as '芯片ID' FROM t_modulids
WHERE modulid>="{start}" and modulid<="{end}" order by modulid asc'''.format(start=mid_start,end=mid_end)
            logging.info(sql)
            cursor = self.conn.cursor(cursor=pymysql.cursors.DictCursor)
            cursor.execute(sql)
            res = cursor.fetchall()
            logging.info
            wb = Workbook()
            ws = wb.active
            ws.append(list(res[0].keys()))
            for each in res:
                ws.append(list(each.values()))
            ws.column_dimensions['A'].width=25
            ws.column_dimensions['B'].width=50
            time_str = time.strftime('%Y_%m%d_%H%M%S', time.localtime())
            self.file_name = ".\\export\\{para}ID对应关系{time}.xlsx".format(para=self.pc,time=time_str)
            wb.save(filename=self.file_name)
            self.statusBar.showMessage("文件已导出:"+self.file_name)
            self.Button_file.setHidden(False)
            self.Button_file.setStyleSheet("background-color : green")

    def export_data(self):
        logging.info("Export data...")
        wb_data =Workbook()
        ws = wb_data.active
        sql = '''
SELECT d.barcode '条形码',d.pc '批次',
case d.result when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '测试结果',
d.workid'工装ID',s.setnosub '子方案序号',s.errcode "错误代码",
case s.z19_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'RF通信结果',
case s.z21_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'RF功耗结果',
case s.z7_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '功耗结果',
case s.z4_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '版本号测试结果',
case s.z5_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'PLC通信结果',
case s.z6_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '频偏结果',
case s.z8_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '过零结果',
case s.z10_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS '停电上报结果',
case s.z18_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'RF校准结果',
case s.z22_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'RF电压结果',
case s.z3_0 when 1 then '成功' when 2 then '失败' ELSE '未测试' END AS 'PIN和电压结果',
# s.z19_0 'RF通信结果',s.z19_1 'RX_SNR结果',s.z19_2 'RX平坦度结果',s.z19_3 'RX_RSSI结果',s.z19_4 'TX_SNR结果',s.z19_5 'TX平坦度结果',s.z19_6 'RX_RSSI结果',s.z19_7 '通信成功率结果',
s.z19_9 'RX_SNR值',s.z19_10 'RX_RSSI值',s.z19_21 '通信成功率',s.z19_22 'TB_SNR',s.z19_23 'TB_RSSI',
#s.z20_0 '抗邻带干扰测试结果',s.z20_1,s.z20_2,s.z20_3,s.z20_4,s.z20_5,s.z20_6,s.z20_7,s.z20_8,s.z20_9,
s.z21_0 'RF功耗测试',s.z7_3 '静态功耗(mA)',s.z7_4 '动态功耗(mA)',s.z21_5 'PLC发送电流',s.z21_6 'RF发送电流',TRUNCATE((s.z21_6-s.z7_3),2) '功耗差值',s.z21_7 '双发电流',
s.z4_0 '版本号结果',s.z4_1 '软件版本号结果',s.z4_2'硬件版本号结果',s.z4_3 '软件版本号',s.z4_4 '硬件版本号',#s.z4_5,s.z4_6,s.z4_7,s.z4_8,
s.z5_0 '通信测试',s.z5_1 'RX_SNR结果',s.z5_2 'RX_SNR值',s.z5_3 'RX_平坦度结果',s.z5_4 'RX_平坦度最小值',s.z5_5 'RX_平坦度平均值',s.z5_6 'RX_平坦度最大值',s.z5_36 'RX_RSSI结果',s.z5_37 'RX_RSSI数值',s.z5_38 '通信成功率结果',s.z5_39 '通信成功率数值',s.z5_40 'TX_SNR结果',s.z5_41 'TX_SNR值',s.z5_42 'TX_平坦度结果',s.z5_43 'TX_平坦度数值0',s.z5_44 'TX_平坦度数值1',s.z5_45 'TX_平坦度数值2',s.z5_75 'TX_RSSI结果',s.z5_76 'TX_RSSI数值',
s.z6_0 '频偏测试',s.z6_1 '频偏值int32',
s.z7_0 '功耗测试',s.z7_1 '静态结果',s.z7_2 '动态结果',
s.z8_0 '过零测试结果',s.z8_1 '过零A1',s.z8_2 '过零A2',s.z8_3 '过零B1',s.z8_4 '过零B2',s.z8_5 '过零C1',s.z8_6 '过零C2',
s.z10_0 '停电上报',s.z10_1 'POWERLOSS',s.z10_2 'PLUG',s.z10_3 '电容充电结果',s.z10_4 'CAP_Vuint16',s.z10_5 '12_boost_V结果',s.z10_6 '12_boost_Vuint16',s.z18_0 '校正结果',
s.z18_1 '正确频点结果',s.z18_2 '错误频点结果',s.z18_4 '正确频点DELTA值', s.z18_5 '正确频点DELTA值2',s.z18_6 'ext_fvco',s.z18_7 'gain_wb',s.z18_8 'gain_nb',s.z18_9 'rc',s.z18_10 '错误频点DELTA值1',s.z18_11 '错误频点DELTA值2',
s.z22_0 'RF电压测试结果',s.z22_4 '1.5V电压值',s.z22_5 '2.8V电压值',
IFNULL(s.z15_0,'无') as '红外',
#s.z16_0 '拓扑组件结果',s.z16_1 '拓扑组件功耗值',
s.z3_0 'PIN和电压',s.z3_1 'RXD结果',s.z3_2 'RST结果',s.z3_3 'EVENT结果',s.z3_4 'TXD结果',s.z3_5 'TXD高电平值',s.z3_6 'STA结果',s.z3_7 'STA高电平值',s.z3_8 '1.2V结果',s.z3_9 '1.2V电压值',s.z3_10 '3.3V结果',s.z3_11 '3.3V电压值',s.z3_12 'clk_rst',s.z3_13 'pin_rst'
,s.intime '测试时间'
 FROM xjlcdbnew.t_testdata d ,xjlcdbnew.t_testdatasub s
WHERE  d.pc='{pc}'  and d.id=s.id
AND s.setnosub='001'
AND d.tstep=0  ORDER BY s.intime DESC LIMIT 100000
'''.format(pc=self.pc)
        logging.info("SQL:"+sql)
        cursor = self.conn.cursor(cursor=pymysql.cursors.DictCursor)
        cursor.execute(sql)
        res = cursor.fetchall()
        ws.append(list(res[0].keys()))
        for each in res:
            ws.append(list(each.values()))
 
        border = Border(left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin'))
        for row in ws:
            for cell in row:
                cell.border = border
        ls_a = list(res[0].values())
        ls_b = list(res[0].keys())
        for each in range(len(res[0])):
            len_a = len(str(ls_a[each]))
            len_b = len(str(ls_b[each]))
            ws.column_dimensions[get_column_letter(each+1)].width =  (len_a if(len_a>len_b) else len_b)+6
        time_str = time.strftime('%Y_%m%d_%H%M%S', time.localtime())
        file_name = ".\\export\\"+"{pc}_测试数据{time}.xlsx".format(pc=self.pc,time=time_str)
        wb_data.save(filename=file_name)
        self.statusBar.showMessage("数据导出已完成...")

    def open_file(self):
        logging.info(self.file_name)
        os.system(r'explorer /select,%s'%self.file_name)


    def log_init(self):    
        if not(os.path.exists("LogFile")):
            mkdir("LogFile")
        if not(os.path.exists("export")):
            mkdir("export")
        time_str = time.strftime('%Y_%m%d_%H%M%S', time.localtime())
        log_file_name = ".\\LogFile\\"+time_str+".txt"
        logging.basicConfig(filename=log_file_name,format="%(asctime)s %(name)s:%(levelname)s-->%(message)s",level='DEBUG',)

    def select_pc(self):
        self.pc = self.cBox_pc_filter.currentText()
        self.statusBar.showMessage("已选择批次:"+self.pc)
        sql = '''select mkidbeg,mkidend from t_pc where pc="{pc}"'''.format(pc=self.pc)
        cursor = self.conn.cursor()
        try:
            cursor.execute(sql)
            res = cursor.fetchone()
            if res!=None:
                self.Edit_MidStart.setText(res[0])
                self.Edit_MidEnd.setText(res[1])
            else:
                self.Edit_MidStart.setText('NULL')
                self.Edit_MidEnd.setText('NULL')
        except Exception as reason:
            logging.info("已选择批次:"+self.pc+'  '+str(reason))

    def get_database_config(self,config):
        self.cfg["host"] = config["host"]
        self.cfg["user"] = config["user"]
        self.cfg["passwd"] = config["passwd"]
        self.cfg["port"] = config["port"]
        self.cfg["db"] = config["db"]
        self.save_config()


    def save_config(self):
        f = open("config.json",'w',encoding="utf-8")
        json.dump(self.cfg,f,indent=4)

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    window = MyWindowShow()
    window.show()
    sys.exit(app.exec_())
